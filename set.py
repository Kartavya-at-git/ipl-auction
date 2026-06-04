from openpyxl import load_workbook

FILE_PATH = "Dataset.xlsx"

wb = load_workbook(FILE_PATH)
ws = wb.active

headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

# Detect columns
category_col = None
role_col = None

for idx, h in enumerate(headers):

    h_lower = h.lower()

    if h_lower == "category":
        category_col = idx + 1

    if h_lower in ["role", "type"]:
        role_col = idx + 1

if category_col is None:
    raise Exception("Category column not found")

if role_col is None:
    raise Exception("Role column not found")

# Create columns if missing
if "Set No" not in headers:
    ws.cell(1, ws.max_column + 1).value = "Set No"

headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

if "Set Name" not in headers:
    ws.cell(1, ws.max_column + 1).value = "Set Name"

headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

set_no_col = headers.index("Set No") + 1
set_name_col = headers.index("Set Name") + 1

global_set_no = 1

row = 2

while row <= ws.max_row:

    category = str(
        ws.cell(row, category_col).value
    ).strip()

    role = str(
        ws.cell(row, role_col).value
    ).strip()

    start_row = row

    count = 0

    while row <= ws.max_row:

        c = str(
            ws.cell(row, category_col).value
        ).strip()

        r = str(
            ws.cell(row, role_col).value
        ).strip()

        if c != category or r != role:
            break

        count += 1

        if count > 8:
            break

        row += 1

    local_set = 1

    if start_row > 2:

        prev_name = ws.cell(
            start_row - 1,
            set_name_col
        ).value

        if prev_name:
            if (
                category in str(prev_name)
                and role in str(prev_name)
            ):
                try:
                    local_set = int(
                        str(prev_name).split("Set ")[1]
                    ) + 1
                except:
                    pass

    set_name = (
        f"{category} "
        f"{role} "
        f"Set "
        f"{local_set}"
    )

    for r in range(start_row, start_row + count):

        ws.cell(
            r,
            set_no_col
        ).value = global_set_no

        ws.cell(
            r,
            set_name_col
        ).value = set_name

    global_set_no += 1

wb.save(FILE_PATH)

print("Set Numbers Generated")

