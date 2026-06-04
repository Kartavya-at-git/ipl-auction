from openpyxl import load_workbook
from copy import copy

FILE_PATH = "Dataset.xlsx"

wb = load_workbook(FILE_PATH)
ws = wb.active

headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

# Detect columns
category_col = None
role_col = None

for idx, h in enumerate(headers):

    h_lower = h.lower()

    if h_lower == "category":
        category_col = idx

    if h_lower in ["role", "type"]:
        role_col = idx

if category_col is None:
    raise Exception("Category column not found")

if role_col is None:
    raise Exception("Role column not found")

# Read rows + styles
rows = []

for row in ws.iter_rows(min_row=2):

    values = [cell.value for cell in row]

    styles = []

    for cell in row:
        styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection)
        })

    rows.append({
        "values": values,
        "styles": styles
    })

category_order = [
    "Marquee",
    "Capped",
    "Uncapped"
]

role_order = [
    "Batter",
    "Wicketkeeper",
    "All-Rounder",
    "Bowler"
]

# Create groups
groups = {}

for category in category_order:

    groups[category] = {}

    for role in role_order:
        groups[category][role] = []

for row in rows:

    category = str(
        row["values"][category_col]
    ).strip()

    role = str(
        row["values"][role_col]
    ).strip()

    if category in groups and role in groups[category]:
        groups[category][role].append(row)

# IPL-style ordering
sorted_rows = []

for category in category_order:

    pointers = {
        role: 0
        for role in role_order
    }

    while True:

        added = False

        for role in role_order:

            players = groups[category][role]

            start = pointers[role]
            end = start + 8

            chunk = players[start:end]

            if chunk:

                sorted_rows.extend(chunk)

                pointers[role] += len(chunk)

                added = True

        if not added:
            break

# Rewrite rows preserving formatting
for row_num, row_data in enumerate(sorted_rows, start=2):

    for col_num, value in enumerate(
        row_data["values"],
        start=1
    ):

        cell = ws.cell(row=row_num, column=col_num)

        cell.value = value

        style = row_data["styles"][col_num - 1]

        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = style["number_format"]
        cell.protection = copy(style["protection"])

wb.save(FILE_PATH)

print("IPL Style Sort Completed")




